"""Generate cossacks3_reference.xlsx with 8 sheets."""
import sys, json
sys.stdout.reconfigure(encoding="utf-8")
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.insert(0, str(Path(__file__).resolve().parent.parent / "parser"))
from config import OUTPUT_DIR, PLAYABLE_NATIONS

from config import DATA_JSON
DATA_PATH = DATA_JSON
XLSX_PATH = OUTPUT_DIR / "cossacks3_reference.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="DDDDDD")
HEADER_FONT = Font(bold=True)
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def autosize(ws, max_width=42):
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        try:
            length = max(len(str(c.value)) if c.value is not None else 0 for c in column_cells)
        except ValueError:
            length = 8
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(length + 2, 8), max_width)


def write_sheet(wb, name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
    for row in rows:
        ws.append(row)
    ws.freeze_panes = "A2"
    autosize(ws)
    return ws


def main():
    data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    wb = openpyxl.Workbook()
    # Drop default sheet
    wb.remove(wb.active)

    # ----- 0. README (deprecation notice) -----
    ws_readme = wb.create_sheet("README")
    ws_readme["A1"] = "Cossacks 3 — Reference (LEGACY xlsx)"
    ws_readme["A1"].font = openpyxl.styles.Font(bold=True, size=14)
    ws_readme["A3"] = ("⚠ Это устаревший монолитный xlsx. Актуальная структурированная "
                       "справка — в каталоге docs/reference/ (главы 01-06, нации, "
                       "сравнения). Производные расчёты — в output/reports/. "
                       "Файл сохраняется для обратной совместимости.")
    ws_readme["A3"].alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top")
    ws_readme.row_dimensions[3].height = 70
    ws_readme.column_dimensions["A"].width = 100
    ws_readme["A5"] = "Sanity checks: см. лист «Sanity_checks»."

    # ----- 1. Buildings -----
    headers = ["sid", "nation", "name_en", "name_ru", "kind", "cluster",
               "hp", "buildtime_sec", "costpercent",
               "food", "wood", "stone", "gold", "iron", "coal",
               "farm (pop bonus)", "score", "usage_short", "produces",
               "capturable", "vision",
               "weapon_damage", "weapon_pause_sec", "weapon_range_tiles",
               "weapon_kind", "weapon_cost",
               "consume", "produce_per_tick", "peasantabsorber", "resourcebase",
               "raw_usage"]
    rows = []
    for b in sorted(data["buildings"], key=lambda x: (x["nation"], x["kind"], x["sid"])):
        rows.append([
            b["sid"], b["nation"], b["name_en"] or "", b["name_ru"] or "",
            b["kind"], b["cluster"] or "",
            b["hp"], b["buildtime_sec"], b["costpercent"],
            b["food"], b["wood"], b["stone"], b["gold"], b["iron"], b["coal"],
            b["farm"], b["score"],
            b.get("usage_short", ""),
            ", ".join(b.get("produces") or []),
            b["capturable"], b["vision"],
            b["weapon_damage"],
            None if b["weapon_pause_frames"] is None else round(b["weapon_pause_frames"]/32, 2),
            None if b["weapon_radiusmax"] is None else round(b["weapon_radiusmax"]/53.3333, 2),
            b["weapon_kind"] or "",
            json.dumps(b["weapon_cost"], ensure_ascii=False) if b["weapon_cost"] else "",
            json.dumps(b["consume"], ensure_ascii=False) if b["consume"] else "",
            json.dumps(b["produce"], ensure_ascii=False) if b["produce"] else "",
            b["peasantabsorber"] or "",
            ", ".join(b["resourcebase"]) if b["resourcebase"] else "",
            b["usage"] or "",
        ])
    write_sheet(wb, "Buildings", headers, rows)

    # ----- 2. Units -----
    headers = ["sid", "nation", "name_en", "name_ru", "trained_in", "usage_short",
               "uniqueness", "available_in_nations",
               "hp", "buildtime_sec",
               "food", "wood", "stone", "gold", "iron", "coal",
               "score", "vision", "searchradius_tiles",
               "aiforce", "walkintervalfactor", "shield", "speed",
               "prot_pike", "prot_sword", "prot_bullet", "prot_cannister",
               "prot_arrow", "prot_cannonball",
               "weapon0_damage", "weapon0_pause_sec",
               "weapon0_range_min_tiles", "weapon0_range_max_tiles",
               "weapon0_kind", "weapon0_id", "weapon0_cost",
               "weapon1_damage", "weapon1_pause_sec", "weapon1_range_max_tiles", "weapon1_kind",
               "consume", "peasantabsorber",
               "transport (ship cargo)", "fishingmax", "fishingspeed", "raw_usage"]
    rows = []
    for u in sorted(data["units"], key=lambda x: (x["nation"], x["sid"])):
        weapons = u["weapons"] or []
        w0 = weapons[0] if len(weapons) > 0 else {}
        w1 = weapons[1] if len(weapons) > 1 else {}
        rows.append([
            u["sid"], u["nation"], u["name_en"] or "", u["name_ru"] or "",
            ", ".join(u.get("trained_in", []) or []),
            u.get("usage_short", ""),
            u.get("uniqueness", ""), u.get("available_in_nations"),
            u["hp"], u["buildtime_sec"],
            u["food"], u["wood"], u["stone"], u["gold"], u["iron"], u["coal"],
            u["score"], u["vision"], u.get("searchradius_tiles"),
            u["aiforce"], u["walkintervalfactor"] or "", u["shield"] or "", u["speed"] or "",
            u["prot_pike"], u["prot_sword"], u["prot_bullet"], u["prot_cannister"],
            u["prot_arrow"], u["prot_cannonball"],
            w0.get("damage"), w0.get("pause_sec"),
            w0.get("radiusmin_tiles"), w0.get("radiusmax_tiles"),
            w0.get("kind") or "", w0.get("weaponsid") or "",
            json.dumps(w0.get("cost"), ensure_ascii=False) if w0.get("cost") else "",
            w1.get("damage"), w1.get("pause_sec"), w1.get("radiusmax_tiles"),
            w1.get("kind") or "",
            json.dumps(u["consume"], ensure_ascii=False) if u["consume"] else "",
            u["peasantabsorber"] or "",
            u.get("transport") or "",
            u.get("fishingmax") or "",
            u.get("fishingspeed") or "",
            u["usage"] or "",
        ])
    write_sheet(wb, "Units", headers, rows)

    # ----- 2b. Ships (filtered subset of Units, with shipping-relevant columns) -----
    SHIP_USAGES = {"gc_obj_usage_fisher", "gc_obj_usage_yacht", "gc_obj_usage_galley",
                    "gc_obj_usage_port", "gc_obj_usage_frigate", "gc_obj_usage_xebec",
                    "gc_obj_usage_battleship", "gc_obj_usage_chaika", "gc_obj_usage_brigantine",
                    "gc_obj_usage_galleon"}
    SHIP_SIDS = {"fishboat", "yacht", "yachttur", "galley", "frigate", "xebec",
                  "battleship", "chaika", "brigantine", "galleon", "sloop"}
    headers = ["sid", "nation", "name_en", "trained_in", "hp", "buildtime_sec",
               "wood", "stone", "gold", "iron", "coal",
               "weapon0_damage", "weapon0_pause_sec", "weapon0_range_max_tiles",
               "weapon0_kind", "weapon0_cost",
               "weapon1_damage", "weapon1_pause_sec", "weapon1_range_max_tiles", "weapon1_kind",
               "vision", "transport", "fishingmax", "fishingspeed",
               "consume_gold (upkeep)", "shield"]
    rows = []
    for u in sorted(data["units"], key=lambda x: (x["sid"], x["nation"])):
        is_ship = (u["sid"] in SHIP_SIDS) or (u.get("usage") in SHIP_USAGES)
        if not is_ship:
            continue
        weapons = u["weapons"] or []
        w0 = weapons[0] if len(weapons) > 0 else {}
        w1 = weapons[1] if len(weapons) > 1 else {}
        consume = u.get("consume") or {}
        rows.append([
            u["sid"], u["nation"], u["name_en"] or "",
            ", ".join(u.get("trained_in", []) or []),
            u["hp"], u["buildtime_sec"],
            u["wood"], u["stone"], u["gold"], u["iron"], u["coal"],
            w0.get("damage"), w0.get("pause_sec"), w0.get("radiusmax_tiles"),
            w0.get("kind") or "",
            json.dumps(w0.get("cost"), ensure_ascii=False) if w0.get("cost") else "",
            w1.get("damage"), w1.get("pause_sec"), w1.get("radiusmax_tiles"),
            w1.get("kind") or "",
            u["vision"], u.get("transport"), u.get("fishingmax"), u.get("fishingspeed"),
            consume.get("gold") if consume else "",
            u.get("shield"),
        ])
    write_sheet(wb, "Ships", headers, rows)

    # ----- 3. Upgrades -----
    headers = ["sid", "nation", "name_en", "name_ru",
               "itype_short", "itype_desc", "value",
               "level", "time_sec",
               "food", "wood", "stone", "gold", "iron", "coal",
               "place", "member", "raw_itype", "tooltiptype", "_source"]
    rows = []
    for u in sorted(data["upgrades"], key=lambda x: (x["nation"], x["sid"])):
        rows.append([
            u["sid"], u["nation"], u["name_en"] or "", u["name_ru"] or "",
            u.get("itype_short") or "", u.get("itype_desc") or "", u["value"],
            u["level"], u["time_sec"],
            u["food"], u["wood"], u["stone"], u["gold"], u["iron"], u["coal"],
            u.get("place") or "", u.get("member") or "", u.get("itype") or "",
            u["tooltiptype"], u.get("_source") or "",
        ])
    write_sheet(wb, "Upgrades", headers, rows)

    # ----- 4. Nations -----
    headers = ["sid", "name_en", "name_ru", "common_cluster",
               "members_count", "upgrade_count", "members_sample"]
    from config import _commonname
    rows = []
    for n in data["nations"]:
        members = n["members"]
        rows.append([
            n["sid"], (n["name_en"] or ""), (n["name_ru"] or ""),
            _commonname(n["sid"]),
            len(members), n["upgrade_count"],
            ", ".join(members[:30]),
        ])
    write_sheet(wb, "Nations", headers, rows)

    # ----- 5. Economy -----
    headers = ["key", "value", "explanation"]
    e = data["economy"]
    rows = [
        ["gc_time_to_frames", e["time_to_frames"], "Frames per second of game time"],
        ["gamespeed_slow",   e["gamespeed_slow"],   "Ticks/sec at game speed 0"],
        ["gamespeed_normal", e["gamespeed_normal"], "Ticks/sec at game speed 1 (default)"],
        ["gamespeed_fast",   e["gamespeed_fast"],   "Ticks/sec at game speed 2"],
        ["max_obj_count",    e["max_obj_count"],    "Hard limit on objects in game"],
        ["max_player_count", e["max_player_count"], "Max players per match"],
        ["field_max_hp",     e["field_max_hp"],     "Field 'HP' (durability per harvest)"],
        ["resource_portion_food",  e["resource_portion_food"],
         "Base food delivered per peasant trip"],
        ["resource_portion_wood",  e["resource_portion_wood"],  "Base wood per peasant trip"],
        ["resource_portion_stone", e["resource_portion_stone"], "Base stone per peasant trip"],
        ["resource_portion_others", e["resource_portion_others"],
         "Base portion for gold/iron/coal/fish (hardcoded in unit.script)"],
        ["hits_needed_food",  e["hits_needed_food"],
         "Peasant 'hits' (work cycles) before delivering food"],
        ["hits_needed_wood",  e["hits_needed_wood"],  "Hits before delivering wood"],
        ["hits_needed_stone", e["hits_needed_stone"], "Hits before delivering stone"],
        ["food_per_unit_upkeep", e["food_per_unit_upkeep"],
         "Food consumed per unit (upkeep)"],
        ["default_eff_percent", e["default_eff_percent"],
         "Default extraction efficiency (player.script:109). Upgrades add to this."],
        ["extraction_formula", "delivered = (base_portion * eff) / 100",
         "Integer division. eff defaults 100, upgrades sum additively."],
    ]
    # Per-resource hourly rate at base eff (1 peasant)
    rows.append(["", "", ""])
    rows.append(["--- Calculated per-peasant rates ---", "", "(at default eff=100, gamespeed=1)"])
    rows.append(["food rate (units/sec, base only)",
                 round(e["resource_portion_food"] / max(e["hits_needed_food"], 1) / 1, 3),
                 "Approximate; ignores travel time. Real: ~45 / (22 hits × 0.5s/hit + travel)."])
    rows.append(["", "", ""])
    rows.append(["--- Damage / protection ---", "", ""])
    rows.append(["pixels_to_tile", e["pixels_to_tile"], "Convert pixels to tiles"])
    rows.append(["damage_formula", e["damage_formula"], "From miscext2.script:_misc_DoDamage"])
    rows.append(["", "", ""])
    rows.append(["--- Object base speeds (abstract units, NOT tiles/sec) ---", "", ""])
    for k, v in e["obj_speed_table_abstract_units"].items():
        rows.append([f"obj_speed_{k}", v,
                       "Relative speed indicator. Empirical test needed for tiles/sec."])
    rows.append(["speed_table_note", e["obj_speed_table_note"], ""])
    write_sheet(wb, "Economy", headers, rows)

    # ----- 6. Combat costs -----
    headers = ["unit_or_building_sid", "nation", "weapon_index", "weapon_kind",
               "damage", "pause_sec", "shots_per_min",
               "iron_per_shot", "coal_per_shot", "gold_per_shot"]
    rows = []
    for u in data["units"]:
        for w in (u["weapons"] or []):
            cost = w.get("cost") or {}
            if cost or w.get("damage"):
                shots = round(60 / w["pause_sec"], 1) if w.get("pause_sec") else None
                rows.append([
                    u["sid"], u["nation"], w["index"], w["kind"] or "",
                    w["damage"], w["pause_sec"], shots,
                    cost.get("iron"), cost.get("coal"), cost.get("gold"),
                ])
    for b in data["buildings"]:
        if b["weapon_damage"] is None:
            continue
        cost = b.get("weapon_cost") or {}
        shots = (round(60 / (b["weapon_pause_frames"]/32), 1)
                 if b["weapon_pause_frames"] else None)
        pause_sec = (round(b["weapon_pause_frames"]/32, 2) if b["weapon_pause_frames"] else None)
        rows.append([
            b["sid"], b["nation"], 0, b.get("weapon_kind") or "",
            b["weapon_damage"], pause_sec, shots,
            cost.get("iron"), cost.get("coal"), cost.get("gold"),
        ])
    write_sheet(wb, "Combat_costs", headers, rows)

    # ----- Officers -----
    headers = ["nation", "officersid", "drummersid", "units_count", "units_list", "call"]
    rows = []
    for o in sorted(data.get("officers", []), key=lambda x: (x["nation"], x["officersid"])):
        rows.append([
            o["nation"], o["officersid"], o["drummersid"], len(o.get("units", [])),
            ", ".join(o.get("units", [])), o.get("call", ""),
        ])
    write_sheet(wb, "Officers", headers, rows)

    # ----- Market rates -----
    headers = ["resource", "buy_min", "buy_default", "buy_max",
                "sell_min", "sell_default", "sell_max",
                "default_buy/sell_ratio"]
    rows = []
    for res, vals in data.get("market_rates", {}).items():
        if res.startswith("_"): continue
        bd = vals["buycostdef"]
        sd = vals["sellcostdef"]
        ratio = round(sd / bd, 3) if bd else "—"
        rows.append([
            res, vals["buycostmin"], bd, vals["buycostmax"],
            round(vals["sellcostmin"], 2), round(sd, 2), round(vals["sellcostmax"], 2),
            ratio,
        ])
    rows.append(["", "", "", "", "", "", "", ""])
    rows.append([data.get("market_rates", {}).get("_note", ""), "", "", "", "", "", "", ""])
    write_sheet(wb, "Market_rates", headers, rows)

    # ----- 7a. Discrepancies (real differences from user's prompt-time notes) -----
    headers = ["fact", "user_note", "file_value", "source", "verdict"]
    rows = []
    for d in data.get("discrepancies", []):
        rows.append([d["fact"], d["user_note"], d["file_value"], d["source"], d["verdict"]])
    write_sheet(wb, "Discrepancies", headers, rows)

    # ----- 7b. Sanity checks (auto-assertions; flags regressions after game patch) -----
    headers = ["category", "check", "expected", "actual", "status"]
    rows = []
    for c in data.get("sanity_checks", []):
        rows.append([
            c["category"], c["name"], str(c["expected"]), str(c["actual"]),
            "PASS" if c["pass"] else "FAIL",
        ])
    ws = write_sheet(wb, "Sanity_checks", headers, rows)
    if rows:
        status_col = chr(ord("A") + 4)
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{len(rows)+1}",
            CellIsRule(operator="equal", formula=['"FAIL"'],
                        fill=PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")),
        )
        ws.conditional_formatting.add(
            f"{status_col}2:{status_col}{len(rows)+1}",
            CellIsRule(operator="equal", formula=['"PASS"'],
                        fill=PatternFill(start_color="DDFFDD", end_color="DDFFDD", fill_type="solid")),
        )

    # ----- 8. Gaps -----
    headers = ["gap", "count", "sample"]
    rows = []
    for g in data["gaps"]:
        rows.append([g["gap"], g["count"], ", ".join(g["sample"][:50])])
    rows.append([
        "Per-unit blacksmith/stable/barracks upgrades",
        "many",
        "These are dynamically created in _country_InitUnitsUpgrades using "
        "loop variables (member, upgplace). The locale catalog (Upgrades sheet) "
        "lists every player-facing upgrade KEY, but cost/time may be missing for "
        "those generated inside loops. See country.script:_country_InitUnitsUpgrades "
        "for the source-of-truth definitions.",
    ])
    rows.append([
        "DLC-only units",
        "small",
        "Some sids in unit.script (e.g., dragoonpol, hackapell) are added to "
        "specific nations in country.script:2904+. They appear in roster correctly.",
    ])
    rows.append([
        "AI/bot-only attributes",
        "—",
        "objprop.aiforce, objprop.bstandground etc. are AI-tuning fields not exposed in xlsx.",
    ])
    write_sheet(wb, "Gaps", headers, rows)

    wb.save(XLSX_PATH)
    print(f"Wrote {XLSX_PATH} ({XLSX_PATH.stat().st_size:,} bytes)")
    print(f"  Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    main()
